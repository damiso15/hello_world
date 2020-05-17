# print('Sonaike Oluwadamilola')
# print('*' * 10)
# price = 10
# price = 20
# print(price)d

# An Exercise to define a variable
# name = 'John Smith'
# age = 20
# is_new = True

# Function. input and print are functions in Python and they have () in front of them
# name = input('Please what is your name? ')
# print('Hi ' + name)

# Exercise
# name = input('What is your name? ')
# colour = input('What is your favourite colour? ')
# print(name + ' likes ' + colour)

# Python Program that will ask for our birth year and calc our age
# birth_year = int(input('birth year: '))
# age = 2020 - birth_year
# print('Your age is: ', age)
# You can use type() to know the type of class a value belongs to.

# Exercise
# user_weight_in_lbs = float(input('Please input your weight: '))
# weight_in_kg = user_weight_in_lbs * 0.4536
# print('Your weight in kg is: %.4fkg' %weight_in_kg)



# Strings in Python
# course = 'Python for Beginners'
# print(course)
# print(course[0])
# print(course[4])
# print(course[-1])
# print(course[-3])
# print(course[0:3])
# print(course[0:])
# print(course[:7])



# Formatted Strings
# first = 'damy'
# last = 'Sonaike'
# message = first + ' [' + last + '] is a coder'
# msg = f'{first} [{last}] is a coder'
# print(msg)



# String Methods
# course = 'Python for Beginners'
# print (len(course)) # It is used to know the length of Characters in a string and can be used to enforce number of inputted characters. It is a general purpose function.
# You use '.' to access other functions that are specific to strings, which we call 'methods' because they belong to strings. There about 91 of them in Python. They are correctly referred to as Methods.
# print(course.upper())
# print(course.lower())
# print(course)
# print(course.find('o')) # We can use this method to find a word. It will return the index of the first occurrence of that letter.
# print(course.replace('Beginners', 'Absolute Beginners')) # We can use this method to replace a letter or word in a string.
# print('Python' in course) # This is an expression that produces a Boolean value. It basically says 'Is the string IN the variable.' This is an operator.



# Arithmetic Operations
# print (10 / 3) # Normal division sign, forward slash. Returns the float value of the operation.
# print(10 // 3) # Double forward slash. Returns the whole number or integer of the operation.
# print(10 % 3) # Modulus, a percent sign, which returns the remainder of the division.
# print(10 ** 3) # Exponential operator.



# Operator Precedence
# Parenthesis always takes priority. Exponential comes first. Division or Multiplication comes next. Finally Addition or Substracation comes last.



# Math Functions
# x = -2.9
# print(round(x)) # This methods round up a float to the nearest integer.
# print(abs(x)) # This is the absolute method that returns the positive value of the given value
# In python we have a Math module function that solves complex mathematically operation but we wil need to import it.
# import math
# print(math.ceil(x))
# print(math.floor(x))



# If Statements
# Look for the conditions.txt file in the 'HelloWorld' venv to continue
# is_hot = False
# is_cold = True

# if is_hot:
# print("It's a hot day")
# print("Drink plenty of water")
# elif is_cold:
# print("It's a cold day")
# print("Wear warm clothes")
# else:
# print("It's a lovely day")

# print("Enjoy your day")

# price = 1000000
# have_good_credit = False

# if have_good_credit:
# down_payment = 0.1 * price
# else:
# down_payment = 0.2 * price

# print(f"Down Payment: ${down_payment}")

# Logical Operators
# has_high_income = True
# has_good_credit = True

# if has_high_income and has_good_credit:
# print("Eligible for loan")

# has_high_income = False
# has_good_credit = True

# if has_high_income or has_good_credit:
# print("Eligible for loan")

# has_good_credit = True
# has_criminal_record = False

# if has_good_credit and not has_criminal_record:
# print("Eligible for loan")



# Comaprison Operators
# we use this when we want to compare a variable with a value

# temperature = int(input("Please enter the temperature: "))

# if temperature > 30:
# print("It's a hot day")
# elif temperature < 10:
# print("It's a cold day")
# else:
# print("It's neither a hot or cold day")

# print("Have a Nice day")

# Exercise
# name = input("Please enter your name: ")

# if len(name) < 3:
# print("ERROR! Name must be at least 3 characters")
# elif len(name) > 50:
# print("ERROR! Name must be a maximum of 50 characters")
# else:
# print("Name looks good!")

# Weight Conversion Exercise
# weight = float(input("Please Enter your current weight: "))
# unit = input("(L)bs or (K)g: ")
# if unit.upper() == 'L':
# conv_weight = weight * 0.4536
# print(f"You are {conv_weight: .4f} pounds")
# elif unit.upper() == 'K':
# conv_weight = weight * 2.2046
# print(f"You are {conv_weight: .4f} kilos")



# While loops
# It is used for interactive softwares and games. It is usually in the form of 'while condition:"
# i = 1
# while i <= 5:
# print(i)
# i += 1
# print("Done")

# i = 1
# while i <= 5:
# print('*' * i)
# i += 1
# print("Done")

# Guessing Games
# A program of a guessing game
# secret_number = 9          # The right number
# guess_count = 0            # i
# guess_limit = 3
# while guess_count < guess_limit:
# guess = int(input("Guess: "))
# guess_count += 1
# if guess == secret_number:
# print("You Won!")
# break
# else:
# print("Sorry, you failed!")

# Exercise A Car Game
# command = ""
# started = False
# while True:
# command = input("> ").lower()
# if command == "start":
# if started:
# print("Car is already started...")
# else:
# started = True
# print("Car started...")
# elif command == "stop":
# if not started:
# print("Car is alreaddy  stopped...")
# else:
# started = False
# print("Car stopped...")
# elif command == "help":
# print("""
# start - to start the car
# stop - to stop te car
# quit - to quit the game
# """)
# elif command == "quit":
# break
# else:
# print("Sorry, I don't understand.")



# For Loops
# We use For Loop to iterate over items of collections such as a string, a string is a sequence of characters.
# for item in 'python':
# print(item)

# for item in ['damy', 'ada', 'john']:
# print(item)

# for item in [1, 2, 3, 4]:
# print(item)

# You can use 'range' with For Loop. It stores an object not a list.
# for item in range(10):
# print(item)

# for item in range(5, 10):   # Range of numbers between 5 and 10.
# print(item)

# for item in range(5, 10, 2):   # Range of numbers between 5 and 10 with a step of 2.
# print(item)

# prices = [10, 20, 30]
# total = 0

# for price in prices:
# total += price

# print(f"Total: {total}")



# Nested Loops
# for x in range(4):
# for y in range(3):
# print(f"({x}, {y})")

# Exercise
# numbers = [5, 2, 5, 2, 2]
# for item in numbers:
# print('x' * item)

# numbers = [5, 2, 5, 2, 2]
# for item in numbers:
# output = ''
# for count in range(item):
# output += 'x'
# print(output)



# Lists
# names = ['John', 'Bob', 'Mosh', 'Sarah', 'Mary']
# print(names)       # print out the strings in the list 'names'.
# print(names[0])    # print out the first index in the list 'names'.
# print(names[0], names[4])
# print(names[-1])   # print out the last index in the list 'names'.
# print(names[0:4])  # print out the indexes of the strings in the list 'names' but won't include the last index.
# print(names[0:])   # Print out all the strings by their indexes in the list 'names'.
# names[0] = 'Jon'   # Changes the first index string
# print(names)

# Exercise - A program to find the largest number in a list.
# numbers = [5, 3, 7, 9, 2, 0, 8, 1, 4, 6]
# max = numbers[0]
# for item in numbers:
# if item > max:
# max = item
# print(max)



# 2D Lists
# matrix = [
# [1, 2, 3],
# [4, 5, 6],
# [7, 8, 9]
# ]
# print(matrix[0])     # Prints the index of the outer list.
# matrix[0][1] = 20    # To change the value of the index of the inner list to 20.
# print(matrix[0][1])  # Prints the index of the inner list.
# for row in matrix:
# for item in row:
# print(item)



# List Methods or Functions
# numbers = [5, 2, 1, 7, 4]
# numbers.append(20)
# print(numbers)
# numbers.insert(0, 10)
# print(numbers)
# numbers.remove(5)
# print(numbers)
# numbers.clear()
# print(numbers)
# numbers = [5, 2, 1, 7, 4]
# numbers.pop()
# print(numbers)
# print(numbers.index(5))
# print(5 in numbers)
# print(numbers.count(5))
# numbers.sort()
# print(numbers)
# numbers.reverse()
# print(numbers)
# numbers2 = numbers.copy()
# print(numbers2)

# Exercise - Write a program to remove the duplicates in a list.
# numbers = [2, 4, 5, 8, 2, 9, 0, 1, 5, 8, 4, 2, 0]
# numbers2 = []
# for item in numbers:
#     if item not in numbers2:
#         numbers2.append(item)
# print(numbers2)



# Tuples
# Tuples are similar to list, they can be used to store a list of items but unlike lists, they cannot me modified.
# We can only get information about Tuples, we cannot change them.
# numbers = (1, 2, 3)
# print(numbers[0])



# Unpacking
# coordinates = (1, 2, 3)
# x, y, z = coordinates
# print(x)
# print(y)
# print(z)



# Dictionaries
# Makes use of Keyvalue:Pairs.
# customer = {
# "name": "John Smith",
# "age": 30,
# "is_verified": True
# }
# print(customer["name"])
# print(customer.get("name"))  # You can use this if you don't want to receive an error messgae.
# print(customer.get("birthedate", "Jan 1 1980"))
# customer["name"] = "Jack Smith"
# print(customer["name"])
# customer["birthdate"] = "Jan 1 1980"
# print(customer)

# Exercise - Converts digits to words.
# phone = input("Phone: ")
# def_digits = {"0": "Zero", "1": "One", "2": "Two", "3": "Three", "4": "Four", "5": "Five", "6": "Six", "7": "Seven", "8": "Eight", "9":"Nine"}
# digit = ""
# for ch in phone:
# digit += def_digits.get(ch, "!") + " "
# print(digit)

# Emoji Converter
# message = input(">")
# words = message.split(' ')
# output = ""
# emojis = {":)": "😁", ":(": "😔"}
# for word in words:
#     output += emojis.get(word, word) + " "
# print(output)

# Break down large piece of code into meanigful and reusable pieces called Functions.
# def greet_user():
# print("Hi there!")
# print('Welcome aboard')


# print("start")
# greet_user()
# print("Finish")



# Parameters
# Parameters in functions are place holders for receiveing informations.
# Arguments in programming are the values we supply to a function.
# def greet_user(name):     # 'name' here is a parameter, which will serve as a local variable to the name we will
# input when calling the function.
# print(f'Hi {name}!')
# print("Welcome aboard")

# print('Start')
# greet_user('damy')
# print('Finish')

# Keyword Arguments
# def greet_user(first_name, last_name):
# print(f'Hi {first_name} {last_name}!')
# print("Welcome aboard")

# print('Start')
# greet_user('damy', 'Sonaike')
# print('Finish')



# Return Statement
# def square(number):
# return number * number

# result = square(3)
# print(result)

# Creating a Reusable Function


# def emoji_converter(message):

# words = message.split(' ')
# emojis = {":)": "😁", ":(": "😔"}
# emoji_output = ""
# for word in words:
# emoji_output += emojis.get(word, word) + " "

# return emoji_output


# message = input("> ")
# result = emoji_converter(message)
# print(result)



# Exceptions
# try:
# age = int(input('Age: '))
# income = 20000
# risk = income / age
# print(age)

# except ZeroDivisionError:
# print('Age cannot be 0')
# except ValueError:
# print('Invalid Value')



# Classes
# We use classes to define new types.
# An object is an instance of a class.
# The class simply defines the blueprint or template for creating objects.
# Objects are the actual instances based on that blueprint.
# Apart from methods this object can also have attributes.
# They are like variables that belongs to a particular object.
# class Point:
# def move(self):
# print("move")

# def draw(self):
# print("draw")


# point = Point()
# point.x = 10
# point.y = 20
# print(point.x)
# point.draw()

# point2 = Point()
# point2.x = 1
# print(point2.x)



# Constructors
# A constructor is a function that gets called at the time of creating an object.
# class Point:
# def __init__(self, x, y):      # Which stands for initialise. 'self' is a reference to the current object
# self.x = x
# self.y = y

# def move(self):
# print("move")

# def draw(self):
# print("draw")


# point = Point(10, 20)
# print(point.x)

# Exercise
# class Person:
# def __init__(self, name):
# self.name = name

# def talk(self):
# print(f"Hi! I am {self.name}")


# damy = Person('Damy')
# damy.talk()

# ada = Person ('Ada')
# ada.talk()



# Inheritance
# class Mammal:
# def walk(self):
# print('walk')


# class Dog(Mammal):
# pass


# class Cat(Mammal):
# pass


# dog = Dog()
# dog.walk()



# Modules
# We use modules  to better organise our codes, instead of writing all the codes and functions in one file.
# We break up our code across multiple files nd it should contain all the related functions and classes.
# Then we can import our module into another module.
# There are two ways to import a module.
# We can import the entire module. This way we get an object with the same name as the module.
# import converters

# print(converters.kg_to_lbs(100))

# The other approach is to import a specific function or class from a module using this method.
# from converters import kg_to_lbs    # Press the keys control and space.

# print(kg_to_lbs(100))

# Exercise

# import utils

# result = utils.find_max()
# print(result)



# Packages
# A package is a container for multiple modules.

# import ecommerce.shippping

# ecommerce.shippping.calc_shipping()

# from ecommerce.shippping import calc_shipping
#
# calc_shipping()
#
# from ecommerce import shippping
#
# shippping.calc_shipping()



# Generating Random Values

# import random

# for i in range(3):
   # print(random.random())          # This function will print values between 0 and 1
   #  print(random.randint(10,20))   # Get random values between 10 to 20

# members = ['John', 'Bob', 'Mary', 'Mosh']
# print(random.choice(members))       # Randomly picks an item from a list and returns it

# Exercise - A program to get 2 random value of a dice.
# import random
#
#
# class Dice:
#     def roll(self):
#
#         first = random.randint(1, 6)
#         second = random.randint(1, 6)
#
#         return first, second
#
#
# ran = Dice()
# print(ran.roll())



# File and Directories
from pathlib import Path

# Now we need to create a Path object to reference a file or a directory on our computer.
# We can use and Absolute Path = start from root of hard disk or a Relative Path = A path starting from the current directory
# Absolute Path = c:\Program Files\Microsoft (For windows)
# Absolute Path = /usr/local/bin (For linux)
# path = Path("ecommerce")
# print(path.exists())
# This path creates a new directory
# path = Path("emails")
# print(path.mkdir())
# This will delete a directory
# path = Path("emails")
# print(path.rmdir())
# We can also find all the files and directory in a given Paths. That is very useful if you want to write a littel program to automate
# something. For example you can iterate over a spreadsheet in a directory, open them and process them.
# path = Path()        # This changes the Path to the current directory.
# print(path.glob("*"))   # With this method we can search for files and directory in the current paths. At the first argument we need to
# pass a string that defines a search pattern. If you type an '*', it means everything, all files and directories. To get all the
# files you use '*.*'. with this pattern we only get the files in the current directory but not the directories.
# for file in path.glob("*"):
#     print(file)



# Pypi and Pip
# Python Package Index (pypi.org) and Package Install Python
# Websscrapping is an engine that browse a website and extracts information for Html files. Browser automation to automate testing
# of web applications. You can use Selenium. To install packages, go to pypi.org, search for the package you are looking for.



# Automating Excel spreadsheet
import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']                                    # Make sure your Sheet starts with capital S.
    cell = sheet['a1']                                      # Method 1 to get the cell of a sheet
    cell = sheet.cell(1, 1)                                 # Method 2 to get the cell of a sheet
    # print(cell.value)
    # print(sheet.max_row)                                  # To get the number of rows in a sheet
    for row in range(2, sheet.max_row + 1):
        # print(row)
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'f2')

    wb.save(filename)



# What is Machine Learning
# Libraries used in Machine Learning include - Numpy, which provides a multidimensional array;
# Pandas, which is a data analysis library that provides a concept called Data Frame;
# MatPlotLib, which is a 2D plotting library for creating graphs and plots;
# Scikit Learn, provides common algoritthms like decision trees, neuron algorithm, etc. The environment used is called Jupyter.
# Create an account on kaggle.com, a website for machine learning.
# Steps in a Machine Learning project:
# 1. Import the Data
# 2. Prepare and Clean the Data
# 3. Split the Data into Training/Test Sets
# 4. Select a Machine Learning Algorithm to build a model
# 5. Train the Model
# 6. Make Predictions
# 7. Evaluate our Algorithm to see it's Accuracy. If it not Accurate we either Fine Tune our Model or Select a Different Algorithm
# from sklearn.tree import DecisionTreeClassifier is a package that comes with Scikit Learn library, which is the most people
# machine library in Python
